from fastapi import APIRouter, Depends, HTTPException, UploadFile, File

from database import supabase
from core import get_admin, hash_password, normalize_email, parse_status, get_optional_user, verify_password, validar_cpf, normalize_cpf
from schemas import UsuarioCreate, UsuarioUpdate, BatchIds, BatchStatus
from routers.ano_letivo import get_ano_letivo_atual
import io
import openpyxl
import csv
from pydantic import BaseModel
from typing import Optional as Opt

def _validar_campos_obrigatorios(valores: dict, campos: list[tuple[str, str]]):
    for chave, rotulo in campos:
        valor = valores.get(chave)
        if valor is None or (isinstance(valor, str) and not valor.strip()):
            raise HTTPException(status_code=400, detail=f'O campo "{rotulo}" não pode ficar em branco.')


router = APIRouter()


def _buscar_conflito_usuario(email: str, ra: str = None, cpf: str = None, excluir_id: int = None):
    filtros = [f"usuEmail.eq.{email}"]
    if ra:
        filtros.append(f"usuRA.eq.{ra}")
    if cpf:
        filtros.append(f"usuCPF.eq.{cpf}")

    query = supabase.table("Usuario").select("*").or_(",".join(filtros))
    if excluir_id:
        query = query.neq("idUsuario", excluir_id)

    encontrados = query.execute().data or []
    if not encontrados:
        return None

    ids_distintos = {u["idUsuario"] for u in encontrados}
    if len(ids_distintos) > 1:
        raise HTTPException(
            status_code=400,
            detail="O e-mail e o RA/CPF informados pertencem a cadastros diferentes. Verifique os dados.",
        )
    return encontrados[0]


def _resposta_conflito(usuario: dict, email: str, ra: str = None, cpf: str = None):
    if usuario.get("usuExcluido") or usuario.get("usuStatus") is not True:
        raise HTTPException(status_code=409, detail="USUARIO_INATIVO")

    if usuario.get("usuEmail") == email:
        campo = "Email"
    elif ra and usuario.get("usuRA") == ra:
        campo = "RA"
    elif cpf and usuario.get("usuCPF") == cpf:
        campo = "CPF"
    else:
        campo = "Cadastro"
    raise HTTPException(status_code=400, detail=f"{campo} já cadastrado")


def _mensagem_conflito_importacao(usuario: dict, email: str, ra: str = None, cpf: str = None) -> str:
    """Traduz um conflito encontrado por _buscar_conflito_usuario em uma mensagem
    amigável para uso nas linhas de erro da importação em massa."""
    if usuario.get("usuExcluido"):
        return "pertence a um cadastro excluído — reative pela tela de cadastro"
    if usuario.get("usuEmail") == email:
        campo = "Email"
    elif ra and usuario.get("usuRA") == ra:
        campo = "RA"
    elif cpf and usuario.get("usuCPF") == cpf:
        campo = "CPF"
    else:
        campo = "Cadastro"
    return f"{campo} já cadastrado"


def _tratar_erro_unicidade(e: Exception, campos: dict):
    msg = str(e)
    if "duplicate key" not in msg and "23505" not in msg:
        return None
    for coluna, rotulo in campos.items():
        if coluna in msg:
            return HTTPException(status_code=409, detail=f"{rotulo} já cadastrado para outro usuário")
    return HTTPException(status_code=409, detail="Já existe um cadastro com esses dados")


def _parse_upload(contents: bytes, filename: str) -> list[dict]:
    """Retorna lista de dicts com as linhas do arquivo (xlsx ou csv)."""
    if filename.lower().endswith(".csv"):
        text = contents.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [
            {k.strip().lower(): (v.strip() if v else "") for k, v in row.items()}
            for row in reader
        ]
    else:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        return rows


# ── ALUNOS ────────────────────────────────────────────────────────

@router.get("/alunos")
def listar_alunos(admin=Depends(get_admin)):
    resp = supabase.table("Usuario").select("*").eq("usuTipo", "Aluno").eq("usuExcluido", False).order("usuNome").execute()
    return resp.data or []


@router.post("/alunos")
def criar_aluno(data: UsuarioCreate, admin=Depends(get_admin)):
    email = normalize_email(data.email)
    ra = (data.ra or "").strip() or None

    _validar_campos_obrigatorios(
        {
            "nome": data.nome,
            "telefone": data.telefone,
            "endereco": data.endereco,
            "ra": ra,
            "serie": data.serie,
        },
        [
            ("nome", "Nome Completo"),
            ("telefone", "Telefone"),
            ("endereco", "Endereço"),
            ("ra", "RA"),
            ("serie", "Série"),
        ],
    )

    email_existe_admin = supabase.table("Administrador").select("*").eq("admEmail", email).execute()
    if email_existe_admin.data:
        raise HTTPException(status_code=400, detail="Email já cadastrado como administrador")

    conflito = _buscar_conflito_usuario(email, ra=ra)
    if conflito:
        _resposta_conflito(conflito, email, ra=ra)

    novo = {
        "usuNome": data.nome,
        "usuEmail": email,
        "usuSenha": hash_password(data.senha),
        "usuTelefone": data.telefone,
        "usuTelefoneResponsavel": data.telefoneResponsavel,
        "usuEndereco": data.endereco,
        "usuRA": ra,
        "usuSerie": data.serie,
        "usuTurma": data.turma,
        "usuAnoLetivo": get_ano_letivo_atual(),
        "usuFormado": False,
        "usuTipo": "Aluno",
        "usuStatus": parse_status(data.status),
        "usuExcluido": False,
    }
    try:
        criado = supabase.table("Usuario").insert(novo).execute()
    except Exception as e:
        erro = _tratar_erro_unicidade(e, {"usuEmail": "Email", "usuRA": "RA"})
        if erro:
            raise erro
        raise HTTPException(status_code=500, detail="Falha ao criar aluno")
    if not criado.data:
        raise HTTPException(status_code=500, detail="Falha ao criar aluno")
    return criado.data[0]


@router.post("/alunos/reativar")
def reativar_aluno(data: UsuarioCreate, admin=Depends(get_admin)):
    email = normalize_email(data.email)
    ra = (data.ra or "").strip() or None

    usuario = _buscar_conflito_usuario(email, ra=ra)
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    payload = {
        "usuNome": data.nome,
        "usuEmail": email,
        "usuSenha": hash_password(data.senha),
        "usuTelefone": data.telefone,
        "usuTelefoneResponsavel": data.telefoneResponsavel,
        "usuEndereco": data.endereco,
        "usuRA": ra,
        "usuSerie": data.serie,
        "usuTurma": data.turma,
        "usuAnoLetivo": get_ano_letivo_atual(),
        "usuFormado": False,
        "usuTipo": "Aluno",
        "usuStatus": True,
        "usuExcluido": False,
    }
    try:
        reativado = supabase.table("Usuario").update(payload).eq("idUsuario", usuario["idUsuario"]).execute()
    except Exception as e:
        erro = _tratar_erro_unicidade(e, {"usuEmail": "Email", "usuRA": "RA"})
        if erro:
            raise erro
        raise HTTPException(status_code=500, detail="Falha ao reativar usuário no banco de dados")
    if not reativado.data:
        raise HTTPException(status_code=500, detail="Falha ao reativar usuário no banco de dados")
    return reativado.data[0]


@router.post("/alunos/importar")
async def importar_alunos(file: UploadFile = File(...), admin=Depends(get_admin)):
    contents = await file.read()
    linhas = _parse_upload(contents, file.filename)
    resultados = {"importados": 0, "ignorados": 0, "erros": []}
    ano_letivo = get_ano_letivo_atual()

    for i, dados in enumerate(linhas, start=2):
        nome  = dados.get("nome", "").strip()
        email = dados.get("email", "").strip().lower()
        ra    = dados.get("ra", "").strip()
        serie = dados.get("serie", "").strip()
        turma = dados.get("turma", "").strip()

        if not nome or not email:
            resultados["erros"].append(f"Linha {i}: nome e email são obrigatórios")
            resultados["ignorados"] += 1
            continue

        if not ra or not serie:
            resultados["erros"].append(f"Linha {i}: RA e Série são obrigatórios para Aluno")
            resultados["ignorados"] += 1
            continue

        try:
            conflito = _buscar_conflito_usuario(email, ra=ra)
        except HTTPException as e:
            resultados["erros"].append(f"Linha {i}: {e.detail}")
            resultados["ignorados"] += 1
            continue
        if conflito:
            resultados["erros"].append(f"Linha {i}: {_mensagem_conflito_importacao(conflito, email, ra=ra)}")
            resultados["ignorados"] += 1
            continue

        novo = {
            "usuNome": nome,
            "usuEmail": email,
            "usuSenha": hash_password("mudar@123"),
            "usuTelefone": dados.get("telefone", ""),
            "usuTelefoneResponsavel": dados.get("telefone_responsavel", ""),
            "usuEndereco": dados.get("endereco", ""),
            "usuRA": ra,
            "usuSerie": serie or None,
            "usuTurma": turma or None,
            "usuAnoLetivo": ano_letivo,
            "usuFormado": False,
            "usuTipo": "Aluno",
            "usuStatus": True,
            "usuExcluido": False,
        }
        try:
            supabase.table("Usuario").insert(novo).execute()
            resultados["importados"] += 1
        except Exception as e:
            resultados["erros"].append(f"Linha {i}: erro ao inserir — {str(e)}")
            resultados["ignorados"] += 1

    return resultados


@router.post("/alunos/batch/excluir")
def excluir_alunos_lote(data: BatchIds, admin=Depends(get_admin)):
    if not data.ids:
        raise HTTPException(status_code=400, detail="Nenhum ID informado")
    resp = (
        supabase.table("Usuario")
        .update({"usuExcluido": True})
        .in_("idUsuario", data.ids)
        .eq("usuTipo", "Aluno")
        .execute()
    )
    afetados = {row["idUsuario"] for row in (resp.data or [])}
    nao_encontrados = [id for id in data.ids if id not in afetados]
    mensagem = f"{len(afetados)} aluno(s) excluído(s) com sucesso"
    if nao_encontrados:
        mensagem += f"; {len(nao_encontrados)} não encontrado(s): {nao_encontrados}"
    return {"message": mensagem, "afetados": len(afetados), "naoEncontrados": nao_encontrados}


@router.post("/alunos/batch/status")
def atualizar_status_lote(data: BatchStatus, admin=Depends(get_admin)):
    if not data.ids:
        raise HTTPException(status_code=400, detail="Nenhum ID informado")
    resp = (
        supabase.table("Usuario")
        .update({"usuStatus": data.status})
        .in_("idUsuario", data.ids)
        .eq("usuTipo", "Aluno")
        .execute()
    )
    afetados = {row["idUsuario"] for row in (resp.data or [])}
    nao_encontrados = [id for id in data.ids if id not in afetados]
    mensagem = f"{len(afetados)} aluno(s) atualizados com sucesso"
    if nao_encontrados:
        mensagem += f"; {len(nao_encontrados)} não encontrado(s): {nao_encontrados}"
    return {"message": mensagem, "afetados": len(afetados), "naoEncontrados": nao_encontrados}

@router.put("/alunos/{idUsuario}")
def atualizar_aluno(idUsuario: int, data: UsuarioUpdate, admin=Depends(get_admin)):
    resp = supabase.table("Usuario").select("*").eq("idUsuario", idUsuario).eq("usuExcluido", False).execute()
    if not resp.data:
        raise HTTPException(status_code=404, detail="Aluno não encontrado")

    payload = {}
    if data.nome is not None:
        payload["usuNome"] = data.nome
    if data.email is not None:
        payload["usuEmail"] = normalize_email(data.email)
    if data.senha is not None:
        payload["usuSenha"] = hash_password(data.senha)
    if data.telefone is not None:
        payload["usuTelefone"] = data.telefone
    if data.telefoneResponsavel is not None:
        payload["usuTelefoneResponsavel"] = data.telefoneResponsavel
    if data.endereco is not None:
        payload["usuEndereco"] = data.endereco
    if data.ra is not None:
        payload["usuRA"] = data.ra
    if data.serie is not None:
        payload["usuSerie"] = data.serie
    if data.turma is not None:
        payload["usuTurma"] = data.turma
    if data.status is not None:
        payload["usuStatus"] = parse_status(data.status)

    if not payload:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")

    mapa_campos = [
        ("usuNome", "nome", "Nome Completo"),
        ("usuTelefone", "telefone", "Telefone"),
        ("usuEndereco", "endereco", "Endereço"),
        ("usuRA", "ra", "RA"),
        ("usuSerie", "serie", "Série"),
    ]
    valores_presentes = {chave: payload[campo] for campo, chave, _ in mapa_campos if campo in payload}
    campos_presentes = [(chave, rotulo) for campo, chave, rotulo in mapa_campos if campo in payload]
    _validar_campos_obrigatorios(valores_presentes, campos_presentes)

    if "usuEmail" in payload or "usuRA" in payload:
        conflito = _buscar_conflito_usuario(
            payload.get("usuEmail", resp.data[0]["usuEmail"]),
            ra=payload.get("usuRA"),
            excluir_id=idUsuario,
        )
        if conflito:
            _resposta_conflito(conflito, payload.get("usuEmail", ""), ra=payload.get("usuRA"))

    try:
        atual = supabase.table("Usuario").update(payload).eq("idUsuario", idUsuario).execute()
    except Exception as e:
        erro = _tratar_erro_unicidade(e, {"usuEmail": "Email", "usuRA": "RA"})
        if erro:
            raise erro
        raise HTTPException(status_code=500, detail="Falha ao atualizar aluno")
    if not atual.data:
        raise HTTPException(status_code=500, detail="Falha ao atualizar aluno")
    return atual.data[0]


@router.delete("/alunos/{idUsuario}")
def deletar_aluno(idUsuario: int, admin=Depends(get_admin)):
    resp = (
        supabase.table("Usuario")
        .update({"usuExcluido": True})
        .eq("idUsuario", idUsuario)
        .eq("usuTipo", "Aluno")
        .execute()
    )
    if not resp.data:
        raise HTTPException(status_code=404, detail="Aluno não encontrado")
    return {"message": "Aluno excluído com sucesso"}

# ── COMUNIDADE ────────────────────────────────────────────────────

@router.get("/comunidade")
def listar_comunidade(admin=Depends(get_admin)):
    resp = supabase.table("Usuario").select("*").eq("usuTipo", "Comunidade").eq("usuExcluido", False).order("usuNome").execute()
    return resp.data or []


@router.post("/comunidade")
def criar_comunidade(data: UsuarioCreate, admin=Depends(get_admin)):
    cpf = normalize_cpf(data.cpf)
    if not validar_cpf(cpf):
        raise HTTPException(status_code=400, detail="CPF inválido")

    email = normalize_email(data.email)

    _validar_campos_obrigatorios(
        {"nome": data.nome, "telefone": data.telefone, "endereco": data.endereco},
        [("nome", "Nome Completo"), ("telefone", "Telefone"), ("endereco", "Endereço")],
    )

    email_existe_admin = supabase.table("Administrador").select("*").eq("admEmail", email).execute()
    if email_existe_admin.data:
        raise HTTPException(status_code=400, detail="Email já cadastrado como administrador")

    conflito = _buscar_conflito_usuario(email, cpf=cpf)
    if conflito:
        _resposta_conflito(conflito, email, cpf=cpf)

    novo = {
        "usuNome": data.nome,
        "usuEmail": email,
        "usuSenha": hash_password(data.senha),
        "usuTelefone": data.telefone,
        "usuTelefoneResponsavel": data.telefoneResponsavel,
        "usuEndereco": data.endereco,
        "usuCPF": cpf,
        "usuTipo": "Comunidade",
        "usuStatus": parse_status(data.status),
        "usuExcluido": False,
    }
    try:
        criado = supabase.table("Usuario").insert(novo).execute()
    except Exception as e:
        erro = _tratar_erro_unicidade(e, {"usuEmail": "Email", "usuCPF": "CPF"})
        if erro:
            raise erro
        raise HTTPException(status_code=500, detail="Falha ao criar membro")
    if not criado.data:
        raise HTTPException(status_code=500, detail="Falha ao criar membro")
    return criado.data[0]


@router.post("/comunidade/reativar")
def reativar_comunidade(data: UsuarioCreate, admin=Depends(get_admin)):
    cpf = normalize_cpf(data.cpf)
    if not validar_cpf(cpf):
        raise HTTPException(status_code=400, detail="CPF inválido")

    email = normalize_email(data.email)

    usuario = _buscar_conflito_usuario(email, cpf=cpf)
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    payload = {
        "usuNome": data.nome,
        "usuEmail": email,
        "usuSenha": hash_password(data.senha),
        "usuTelefone": data.telefone,
        "usuTelefoneResponsavel": data.telefoneResponsavel,
        "usuEndereco": data.endereco,
        "usuCPF": cpf,
        "usuTipo": "Comunidade",
        "usuStatus": True,
        "usuExcluido": False,
        # Ex-aluno retornando como comunidade: limpa os dados acadêmicos
        "usuSerie": None,
        "usuTurma": None,
        "usuAnoLetivo": None,
        "usuFormado": False,
    }
    try:
        reativado = supabase.table("Usuario").update(payload).eq("idUsuario", usuario["idUsuario"]).execute()
    except Exception as e:
        erro = _tratar_erro_unicidade(e, {"usuEmail": "Email", "usuCPF": "CPF"})
        if erro:
            raise erro
        raise HTTPException(status_code=500, detail="Falha ao reativar usuário no banco de dados")
    if not reativado.data:
        raise HTTPException(status_code=500, detail="Falha ao reativar usuário no banco de dados")
    return reativado.data[0]


@router.post("/comunidade/importar")
async def importar_comunidade(file: UploadFile = File(...), admin=Depends(get_admin)):
    contents = await file.read()
    linhas = _parse_upload(contents, file.filename)
    resultados = {"importados": 0, "ignorados": 0, "erros": []}
    cpfs_vistos = set()

    for i, dados in enumerate(linhas, start=2):
        nome  = dados.get("nome", "").strip()
        email = dados.get("email", "").strip().lower()
        cpf   = normalize_cpf(dados.get("cpf", ""))

        if not nome or not email:
            resultados["erros"].append(f"Linha {i}: nome e email são obrigatórios")
            resultados["ignorados"] += 1
            continue

        if cpf and not validar_cpf(cpf):
            resultados["erros"].append(f"Linha {i}: CPF '{cpf}' inválido")
            resultados["ignorados"] += 1
            continue

        if cpf and cpf in cpfs_vistos:
            resultados["erros"].append(f"Linha {i}: CPF '{cpf}' duplicado no arquivo")
            resultados["ignorados"] += 1
            continue

        try:
            conflito = _buscar_conflito_usuario(email, cpf=cpf or None)
        except HTTPException as e:
            resultados["erros"].append(f"Linha {i}: {e.detail}")
            resultados["ignorados"] += 1
            continue
        if conflito:
            resultados["erros"].append(f"Linha {i}: {_mensagem_conflito_importacao(conflito, email, cpf=cpf or None)}")
            resultados["ignorados"] += 1
            continue

        novo = {
            "usuNome": nome,
            "usuEmail": email,
            "usuSenha": hash_password("mudar@123"),
            "usuTelefone": dados.get("telefone", ""),
            "usuTelefoneResponsavel": dados.get("telefone_responsavel", ""),
            "usuEndereco": dados.get("endereco", ""),
            "usuCPF": cpf,
            "usuTipo": "Comunidade",
            "usuStatus": True,
            "usuExcluido": False,
        }
        try:
            supabase.table("Usuario").insert(novo).execute()
            resultados["importados"] += 1
            if cpf:
                cpfs_vistos.add(cpf)
        except Exception as e:
            resultados["erros"].append(f"Linha {i}: erro ao inserir — {str(e)}")
            resultados["ignorados"] += 1

    return resultados


@router.post("/comunidade/batch/excluir")
def excluir_comunidade_lote(data: BatchIds, admin=Depends(get_admin)):
    if not data.ids:
        raise HTTPException(status_code=400, detail="Nenhum ID informado")
    resp = (
        supabase.table("Usuario")
        .update({"usuExcluido": True})
        .in_("idUsuario", data.ids)
        .eq("usuTipo", "Comunidade")
        .execute()
    )
    afetados = {row["idUsuario"] for row in (resp.data or [])}
    nao_encontrados = [id for id in data.ids if id not in afetados]
    mensagem = f"{len(afetados)} membro(s) excluído(s) com sucesso"
    if nao_encontrados:
        mensagem += f"; {len(nao_encontrados)} não encontrado(s): {nao_encontrados}"
    return {"message": mensagem, "afetados": len(afetados), "naoEncontrados": nao_encontrados}


@router.post("/comunidade/batch/status")
def atualizar_status_comunidade_lote(data: BatchStatus, admin=Depends(get_admin)):
    if not data.ids:
        raise HTTPException(status_code=400, detail="Nenhum ID informado")
    resp = (
        supabase.table("Usuario")
        .update({"usuStatus": data.status})
        .in_("idUsuario", data.ids)
        .eq("usuTipo", "Comunidade")
        .execute()
    )
    afetados = {row["idUsuario"] for row in (resp.data or [])}
    nao_encontrados = [id for id in data.ids if id not in afetados]
    mensagem = f"{len(afetados)} membro(s) atualizados com sucesso"
    if nao_encontrados:
        mensagem += f"; {len(nao_encontrados)} não encontrado(s): {nao_encontrados}"
    return {"message": mensagem, "afetados": len(afetados), "naoEncontrados": nao_encontrados}

@router.put("/comunidade/{idUsuario}")
def atualizar_comunidade(idUsuario: int, data: UsuarioUpdate, admin=Depends(get_admin)):
    resp = supabase.table("Usuario").select("*").eq("idUsuario", idUsuario).eq("usuExcluido", False).execute()
    if not resp.data:
        raise HTTPException(status_code=404, detail="Membro não encontrado")

    payload = {}
    if data.nome is not None:
        payload["usuNome"] = data.nome
    if data.email is not None:
        payload["usuEmail"] = normalize_email(data.email)
    if data.senha is not None:
        payload["usuSenha"] = hash_password(data.senha)
    if data.telefone is not None:
        payload["usuTelefone"] = data.telefone
    if data.telefoneResponsavel is not None:
        payload["usuTelefoneResponsavel"] = data.telefoneResponsavel
    if data.endereco is not None:
        payload["usuEndereco"] = data.endereco
    if data.cpf is not None:
        cpf = normalize_cpf(data.cpf)
        if not validar_cpf(cpf):
            raise HTTPException(status_code=400, detail="CPF inválido")
        payload["usuCPF"] = cpf
    if data.status is not None:
        payload["usuStatus"] = parse_status(data.status)

    if not payload:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")

    mapa_campos = [
        ("usuNome", "nome", "Nome Completo"),
        ("usuTelefone", "telefone", "Telefone"),
        ("usuEndereco", "endereco", "Endereço"),
    ]
    valores_presentes = {chave: payload[campo] for campo, chave, _ in mapa_campos if campo in payload}
    campos_presentes = [(chave, rotulo) for campo, chave, rotulo in mapa_campos if campo in payload]
    _validar_campos_obrigatorios(valores_presentes, campos_presentes)

    # Se e-mail ou CPF estão sendo trocados, checa se já pertencem a OUTRO
    # usuário (ativo, inativo ou excluído) antes de tentar salvar.
    if "usuEmail" in payload or "usuCPF" in payload:
        conflito = _buscar_conflito_usuario(
            payload.get("usuEmail", resp.data[0]["usuEmail"]),
            cpf=payload.get("usuCPF"),
            excluir_id=idUsuario,
        )
        if conflito:
            _resposta_conflito(conflito, payload.get("usuEmail", ""), cpf=payload.get("usuCPF"))

    try:
        atual = supabase.table("Usuario").update(payload).eq("idUsuario", idUsuario).execute()
    except Exception as e:
        erro = _tratar_erro_unicidade(e, {"usuEmail": "Email", "usuCPF": "CPF"})
        if erro:
            raise erro
        raise HTTPException(status_code=500, detail="Falha ao atualizar membro")
    if not atual.data:
        raise HTTPException(status_code=500, detail="Falha ao atualizar membro")
    return atual.data[0]


@router.delete("/comunidade/{idUsuario}")
def deletar_comunidade(idUsuario: int, admin=Depends(get_admin)):
    resp = (
        supabase.table("Usuario")
        .update({"usuExcluido": True})
        .eq("idUsuario", idUsuario)
        .eq("usuTipo", "Comunidade")
        .execute()
    )
    if not resp.data:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    return {"message": "Membro da comunidade excluído com sucesso"}


# ── PERFIL DO PRÓPRIO USUÁRIO ─────────────────────────────────────


class PerfilUpdate(BaseModel):
    telefone:             Opt[str] = None
    telefoneResponsavel:  Opt[str] = None
    endereco:             Opt[str] = None
    senhaAtual:           Opt[str] = None
    novaSenha:            Opt[str] = None

@router.get("/usuario/me")
def get_perfil(user=Depends(get_optional_user)):
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")
    resp = supabase.table("Usuario").select("*").eq("usuEmail", user["sub"]).execute()
    if not resp.data:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    u = resp.data[0]
    return {
        "idUsuario":            u.get("idUsuario"),
        "nome":                 u.get("usuNome"),
        "email":                u.get("usuEmail"),
        "ra":                   u.get("usuRA"),
        "cpf":                  u.get("usuCPF"),
        "dataNascimento":       u.get("usuDataNascimento"),
        "endereco":             u.get("usuEndereco"),
        "telefone":             u.get("usuTelefone"),
        "telefoneResponsavel":  u.get("usuTelefoneResponsavel"),
        "tipo":                 u.get("usuTipo"),
    }

@router.patch("/usuario/me")
def atualizar_perfil(data: PerfilUpdate, user=Depends(get_optional_user)):
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")

    resp = supabase.table("Usuario").select("*").eq("usuEmail", user["sub"]).execute()
    if not resp.data:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    u = resp.data[0]

    payload = {}

    # Contato
    if data.telefone is not None:
        payload["usuTelefone"] = data.telefone
    if data.telefoneResponsavel is not None:
        payload["usuTelefoneResponsavel"] = data.telefoneResponsavel
    if data.endereco is not None:
        payload["usuEndereco"] = data.endereco

    # Senha: exige senha atual correta
    if data.novaSenha:
        if not data.senhaAtual:
            raise HTTPException(status_code=400, detail="Informe a senha atual para trocar a senha")
        if not verify_password(data.senhaAtual, u.get("usuSenha", "")):
            raise HTTPException(status_code=400, detail="Senha atual incorreta")
        payload["usuSenha"] = hash_password(data.novaSenha)

    if not payload:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")

    atual = supabase.table("Usuario").update(payload).eq("idUsuario", u["idUsuario"]).execute()
    if not atual.data:
        raise HTTPException(status_code=500, detail="Falha ao atualizar perfil")

    updated = atual.data[0]
    return {
        "idUsuario":            updated.get("idUsuario"),
        "nome":                 updated.get("usuNome"),
        "email":                updated.get("usuEmail"),
        "ra":                   updated.get("usuRA"),
        "cpf":                  updated.get("usuCPF"),
        "dataNascimento":       updated.get("usuDataNascimento"),
        "endereco":             updated.get("usuEndereco"),
        "telefone":             updated.get("usuTelefone"),
        "telefoneResponsavel":  updated.get("usuTelefoneResponsavel"),
        "tipo":                 updated.get("usuTipo"),
    }